"""
Generate the "before" artifact: a messy, manually-maintained monthly P&L
workbook of the kind a controller actually keeps by hand.

It is deliberately human-formatted (title banners, blank spacer rows, business
units stacked in blocks, months pivoted across columns) — exactly the shape
that is painful to analyze and the reason the rest of this pipeline exists.

Run: python data_gen.py  ->  data/raw_pnl.xlsx
"""

from __future__ import annotations

import os
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

OUT_DIR = "data"
N_MONTHS = 24
BUSINESS_UNITS = ["Retail", "Wholesale", "Online", "Services"]
# P&L accounts we store at the granular level; everything else is derived.
ACCOUNTS = ["Revenue", "COGS", "Sales & Marketing", "G&A", "R&D"]

# Per-BU scale and economics (revenue base in BRL '000, margins as fractions).
BU_PROFILE = {
    "Retail":    dict(base=1800, growth=0.010, season=0.18, cogs=0.62, sm=0.10, ga=0.07, rd=0.00),
    "Wholesale": dict(base=2600, growth=0.006, season=0.08, cogs=0.71, sm=0.05, ga=0.05, rd=0.00),
    "Online":    dict(base=900,  growth=0.028, season=0.22, cogs=0.55, sm=0.16, ga=0.06, rd=0.03),
    "Services":  dict(base=700,  growth=0.018, season=0.05, cogs=0.40, sm=0.08, ga=0.10, rd=0.05),
}


def _month_labels(n: int) -> list[str]:
    end = pd.Timestamp.today().normalize().replace(day=1)
    start = end - pd.DateOffset(months=n - 1)
    return [d.strftime("%b/%y") for d in pd.date_range(start, end, freq="MS")]


def _series(profile: dict, rng: np.random.Generator) -> dict[str, np.ndarray]:
    """Monthly Revenue and cost lines for one business unit."""
    months = np.arange(N_MONTHS)
    trend = profile["base"] * (1 + profile["growth"]) ** months
    season = 1 + profile["season"] * np.sin(2 * np.pi * (months % 12) / 12 - 0.5)
    noise = rng.normal(1.0, 0.05, N_MONTHS)
    revenue = trend * season * noise

    return {
        "Revenue": revenue,
        "COGS": revenue * profile["cogs"] * rng.normal(1.0, 0.03, N_MONTHS),
        "Sales & Marketing": revenue * profile["sm"] * rng.normal(1.0, 0.06, N_MONTHS),
        "G&A": revenue * profile["ga"] * rng.normal(1.0, 0.04, N_MONTHS),
        "R&D": revenue * profile["rd"] * rng.normal(1.0, 0.08, N_MONTHS),
    }


def build_actuals(rng: np.random.Generator) -> dict[tuple[str, str], np.ndarray]:
    data = {}
    for bu, profile in BU_PROFILE.items():
        s = _series(profile, rng)
        for acct in ACCOUNTS:
            data[(bu, acct)] = np.round(s[acct], 0)
    return data


def write_ugly_workbook(actuals: dict, months: list[str], path: str) -> None:
    """Write the human-formatted (messy) workbook with openpyxl."""
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L"

    title_font = Font(bold=True, size=14)
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="44546A")
    bu_font = Font(bold=True, italic=True, color="1F3864")

    ws["A1"] = "MERIDIAN INDUSTRIES S.A."
    ws["A1"].font = title_font
    ws["A2"] = "Monthly Management P&L  —  values in BRL '000  —  CONFIDENTIAL"
    ws["A2"].font = Font(italic=True, size=10, color="808080")
    # row 3 left blank on purpose

    header_row = 4
    ws.cell(header_row, 1, "Business Unit").font = hdr_font
    ws.cell(header_row, 2, "Account").font = hdr_font
    for j, m in enumerate(months):
        c = ws.cell(header_row, 3 + j, m)
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center")
    for col in range(1, 3 + len(months)):
        ws.cell(header_row, col).fill = hdr_fill

    r = header_row + 1
    for bu in BUSINESS_UNITS:
        # a BU banner row, then its accounts (BU name only on first line -
        # the kind of "merged-ish" manual layout the engine must clean up)
        first = True
        for acct in ACCOUNTS:
            ws.cell(r, 1, bu if first else None)
            if first:
                ws.cell(r, 1).font = bu_font
            ws.cell(r, 2, acct)
            for j in range(len(months)):
                ws.cell(r, 3 + j, float(actuals[(bu, acct)][j]))
            r += 1
            first = False
        r += 1  # blank spacer row between business units

    # Budget on a second sheet: a single annual budget per BU/account (so the
    # engine has to spread/compare against a different granularity).
    wb_ws2 = wb.create_sheet("Budget FY")
    wb_ws2["A1"] = "Annual Budget (BRL '000)"
    wb_ws2["A1"].font = title_font
    wb_ws2.cell(3, 1, "Business Unit").font = hdr_font
    wb_ws2.cell(3, 2, "Account").font = hdr_font
    wb_ws2.cell(3, 3, "Full-Year Budget").font = hdr_font
    for col in range(1, 4):
        wb_ws2.cell(3, col).fill = hdr_fill
    rr = 4
    for bu in BUSINESS_UNITS:
        for acct in ACCOUNTS:
            # budget = last-12-months actual nudged by a target factor
            last12 = actuals[(bu, acct)][-12:].sum()
            target = last12 * (1.05 if acct == "Revenue" else 0.98)
            wb_ws2.cell(rr, 1, bu)
            wb_ws2.cell(rr, 2, acct)
            wb_ws2.cell(rr, 3, float(round(target, 0)))
            rr += 1

    for ws_ in (ws, wb_ws2):
        ws_.column_dimensions["A"].width = 16
        ws_.column_dimensions["B"].width = 18

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)


def main() -> None:
    rng = np.random.default_rng(2024)
    months = _month_labels(N_MONTHS)
    actuals = build_actuals(rng)
    path = os.path.join(OUT_DIR, "raw_pnl.xlsx")
    write_ugly_workbook(actuals, months, path)
    print(f"Wrote {path}  ({N_MONTHS} months, {len(BUSINESS_UNITS)} business units)")


if __name__ == "__main__":
    main()
